"""Pipeline to fetch member data from Admidio and export to Excel."""

from __future__ import annotations

import subprocess
import tempfile
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from st_andreas.admidio_db import (
    AdmidioField,
    db_connection,
    fetch_field_value_list,
    fetch_user_field_values,
    load_db_config,
    load_secrets,
    load_ssh_config,
    ssh_tunnel,
)

ADMIDIO_VOLUME_PATH: Final = (
    "/var/lib/docker/volumes/"
    "756e80f3bc09b1883b34a1389fce457f3561e7711d4ec592edbda3f2b422ad5a/"
    "_data/documents_sta/Mitgliederliste"
)
ADMIDIO_MITGLIEDERLISTE_FOLDER_ID: Final = 3
ADMIDIO_SYSTEM_USER_ID: Final = 1


@dataclass
class MemberRecord:
    """Member data record."""

    mitglieds_nr: str | None = None
    nachname: str | None = None
    vorname: str | None = None
    kontoinhaber: str | None = None
    familien_nr: str | None = None
    sippe: str | None = None


COLUMN_HEADERS: Final = [
    "MitgliedsNr",
    "Nachname",
    "Vorname",
    "Kontoinhaber",
    "FamilienNr",
    "Sippe",
]

FIELD_TO_ATTR: Final[dict[str, str]] = {
    "MITGLIEDSNR": "mitglieds_nr",
    "LAST_NAME": "nachname",
    "FIRST_NAME": "vorname",
    "KONTOINHABER": "kontoinhaber",
    "FAMILIENNR": "familien_nr",
    "SIPPE": "sippe",
}


def fetch_member_data(
    sippe_mapping: dict[str, str],
    table_prefix: str,
) -> list[MemberRecord]:
    """Fetch member data from Admidio database."""
    field_ids = [
        AdmidioField.MITGLIEDSNR.value,
        AdmidioField.FAMILIENNR.value,
        AdmidioField.SIPPE.value,
        AdmidioField.KONTOINHABER.value,
        AdmidioField.LAST_NAME.value,
        AdmidioField.FIRST_NAME.value,
    ]

    with db_connection() as conn:
        users = fetch_user_field_values(conn, field_ids, table_prefix)

    members: list[MemberRecord] = []
    for user_data in users.values():
        member = MemberRecord()
        for field_name, value in user_data.items():
            if field_name == "SIPPE" and value:
                value = sippe_mapping.get(value, value)

            attr_name = FIELD_TO_ATTR.get(field_name)
            if attr_name:
                setattr(member, attr_name, value)

        members.append(member)

    return members


def export_to_excel(members: list[MemberRecord], output_path: Path) -> None:
    """Export member records to styled Excel file."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sorted_members = sorted(members, key=lambda m: (m.sippe or "", m.nachname or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Mitglieder"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    data_alignment = Alignment(vertical="center")
    alt_row_fill = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
    )

    ws.append(COLUMN_HEADERS)

    for col_num, _ in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_num, member in enumerate(sorted_members, 2):
        row_data = [
            member.mitglieds_nr,
            member.nachname,
            member.vorname,
            member.kontoinhaber,
            member.familien_nr,
            member.sippe,
        ]
        ws.append(row_data)

        for col_num in range(1, len(COLUMN_HEADERS) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = data_alignment
            if row_num % 2 == 0:
                cell.fill = alt_row_fill

    column_widths = [15, 18, 18, 38, 12, 22]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[1].height = 25

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def generate_filename() -> str:
    """Generate filename with current date and time."""
    now = datetime.now().strftime("%Y-%m-%d_%H%M")
    return f"{now}_Mitgliederliste.xlsx"


def upload_to_admidio(
    local_file: Path,
    remote_filename: str,
    table_prefix: str,
) -> None:
    """Upload file to Admidio and register in database."""
    ssh_config = load_ssh_config()
    db_config = load_db_config()

    expanded_key_path = Path(ssh_config.key_path).expanduser()
    remote_path = f"{ADMIDIO_VOLUME_PATH}/{remote_filename}"

    scp_cmd = [
        "scp",
        "-i",
        str(expanded_key_path),
        "-o",
        "StrictHostKeyChecking=no",
        str(local_file),
        f"{ssh_config.user}@{ssh_config.host}:{remote_path}",
    ]
    subprocess.run(scp_cmd, check=True)

    chown_cmd = [
        "ssh",
        "-i",
        str(expanded_key_path),
        "-o",
        "StrictHostKeyChecking=no",
        f"{ssh_config.user}@{ssh_config.host}",
        f"chown www-data:www-data '{remote_path}'",
    ]
    subprocess.run(chown_cmd, check=True)

    file_uuid = str(uuid.uuid4())

    with db_connection(db_config) as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT fil_id FROM {table_prefix}files 
                WHERE fil_fol_id = %s AND fil_name = %s
                """,
                (ADMIDIO_MITGLIEDERLISTE_FOLDER_ID, remote_filename),
            )
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    f"""
                    UPDATE {table_prefix}files 
                    SET fil_timestamp = NOW()
                    WHERE fil_id = %s
                    """,
                    (existing["fil_id"],),
                )
            else:
                cursor.execute(
                    f"""
                    INSERT INTO {table_prefix}files 
                    (fil_fol_id, fil_uuid, fil_name, fil_locked, fil_counter,
                     fil_usr_id, fil_timestamp)
                    VALUES (%s, %s, %s, 0, 0, %s, NOW())
                    """,
                    (
                        ADMIDIO_MITGLIEDERLISTE_FOLDER_ID,
                        file_uuid,
                        remote_filename,
                        ADMIDIO_SYSTEM_USER_ID,
                    ),
                )

        conn.commit()


def main() -> None:
    """Main entry point for the member data pipeline."""
    secrets = load_secrets()
    table_prefix = secrets["ADMIDIO_TABLE_PREFIX"]

    with ssh_tunnel():
        with db_connection() as conn:
            sippe_mapping = fetch_field_value_list(conn, "SIPPE", table_prefix)

        members = fetch_member_data(sippe_mapping, table_prefix)

        with tempfile.TemporaryDirectory() as tmpdir:
            filename = generate_filename()
            local_file = Path(tmpdir) / filename

            export_to_excel(members, local_file)
            upload_to_admidio(local_file, filename, table_prefix)

    print(f"Uploaded {len(members)} members to Admidio as {filename}")


if __name__ == "__main__":
    main()
