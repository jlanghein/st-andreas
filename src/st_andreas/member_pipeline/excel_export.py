"""Excel export utilities for member pipelines."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Final

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from st_andreas.member_pipeline.config import ColumnConfig

HEADER_FONT_COLOR: Final[str] = "FFFFFF"
HEADER_FILL_COLOR: Final[str] = "4472C4"
HEADER_FONT_SIZE: Final[int] = 11
HEADER_ROW_HEIGHT: Final[int] = 25

BORDER_COLOR: Final[str] = "B4B4B4"
ALT_ROW_FILL_COLOR: Final[str] = "D9E2F3"


def export_to_excel(
    df: pd.DataFrame,
    columns: tuple[ColumnConfig, ...],
    output_path: Path,
    sheet_name: str = "Mitglieder",
) -> None:
    """Export DataFrame to styled Excel file."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=HEADER_FONT_SIZE)
    header_fill = PatternFill(
        start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    data_alignment = Alignment(vertical="center")
    alt_row_fill = PatternFill(
        start_color=ALT_ROW_FILL_COLOR, end_color=ALT_ROW_FILL_COLOR, fill_type="solid"
    )

    headers = [col.header for col in columns]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_num, (_, row) in enumerate(df.iterrows(), start=2):
        row_data = [row.get(col.header) for col in columns]
        ws.append(row_data)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = data_alignment
            if row_num % 2 == 0:
                cell.fill = alt_row_fill

    for col_num, col_config in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = col_config.width

    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
