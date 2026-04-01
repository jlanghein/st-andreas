"""Pipeline to fetch member data from Admidio and export to Excel."""

from __future__ import annotations

import subprocess
import tempfile
import time
import uuid
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Final

import pymysql
import pymysql.cursors
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SECRETS_FILE: Final = Path(__file__).parent / "secrets.env"
OUTPUT_DIR: Final = Path(__file__).parent / "data"
OUTPUT_FILE: Final = OUTPUT_DIR / "mitglieder.xlsx"

SSH_TUNNEL_LOCAL_PORT: Final = 13306
SSH_TUNNEL_REMOTE_PORT: Final = 3306
DOCKER_DB_HOST: Final = "172.18.0.2"

ADMIDIO_VOLUME_PATH: Final = (
    "/var/lib/docker/volumes/"
    "756e80f3bc09b1883b34a1389fce457f3561e7711d4ec592edbda3f2b422ad5a/"
    "_data/documents_sta/Mitgliederliste"
)
ADMIDIO_MITGLIEDERLISTE_FOLDER_ID: Final = 3
ADMIDIO_SYSTEM_USER_ID: Final = 1


class AdmidioField(Enum):
    """Admidio user field IDs mapped to their internal names."""

    MITGLIEDSNR = 20
    FAMILIENNR = 21
    SIPPE = 24
    KONTOINHABER = 32
    LAST_NAME = 1
    FIRST_NAME = 2


@dataclass
class MemberRecord:
    """Member data record."""

    mitglieds_nr: str | None = None
    nachname: str | None = None
    vorname: str | None = None
    kontoinhaber: str | None = None
    familien_nr: str | None = None
    sippe: str | None = None


COLUMN_HEADERS: Final = [
    "MitgliedsNr",
    "Nachname",
    "Vorname",
    "Kontoinhaber",
    "FamilienNr",
    "Sippe",
]

FIELD_TO_ATTR: Final[dict[str, str]] = {
    "MITGLIEDSNR": "mitglieds_nr",
    "LAST_NAME": "nachname",
    "FIRST_NAME": "vorname",
    "KONTOINHABER": "kontoinhaber",
    "FAMILIENNR": "familien_nr",
    "SIPPE": "sippe",
}


def load_secrets(path: Path) -> dict[str, str]:
    """Load environment variables from a secrets file."""
    secrets: dict[str, str] = {}
    with path.open() as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, value = line.split("=", 1)
                secrets[key.strip()] = value.strip()
    return secrets


@contextmanager
def ssh_tunnel(
    ssh_host: str,
    ssh_user: str,
    ssh_key_path: str,
    local_port: int,
    remote_host: str,
    remote_port: int,
):
    """Create an SSH tunnel to access remote database."""
    expanded_key_path = Path(ssh_key_path).expanduser()
    cmd = [
        "ssh",
        "-i",
        str(expanded_key_path),
        "-o",
        "StrictHostKeyChecking=no",
        "-o",
        "ExitOnForwardFailure=yes",
        "-L",
        f"{local_port}:{remote_host}:{remote_port}",
        "-N",
        f"{ssh_user}@{ssh_host}",
    ]
    process = subprocess.Popen(
        cmd,
        stdin=subprocess.DEVNULL,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    time.sleep(2)

    if process.poll() is not None:
        raise RuntimeError("SSH tunnel failed to start")

    try:
        yield
    finally:
        process.terminate()
        try:
            process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            process.kill()


def fetch_sippe_mapping(
    host: str,
    port: int,
    database: str,
    user: str,
    password: str,
    table_prefix: str,
) -> dict[str, str]:
    """Fetch Sippe ID to name mapping from database."""
    query = f"""
        SELECT usf_value_list 
        FROM {table_prefix}user_fields 
        WHERE usf_name_intern = 'SIPPE'
    """

    with pymysql.connect(
        host=host,
        port=port,
        database=database,
        user=user,
        password=password,
        cursorclass=pymysql.cursors.DictCursor,
    ) as conn:
        with conn.cursor() as cursor:
            cursor.execute(query)
            result = cursor.fetchone()

    if not result or not result["usf_value_list"]:
        return {}

    sippe_names = result["usf_value_list"].split("\n")
    return {str(i + 1): name.strip() for i, name in enumerate(sippe_names)}


def fetch_member_data(
    host: str,
    port: int,
    database: str,
    user: str,
    password: str,
    table_prefix: str,
    sippe_mapping: dict[str, str],
) -> list[MemberRecord]:
    """Fetch member data from Admidio database."""
    field_ids = [f.value for f in AdmidioField]

    query = f"""
        SELECT
            u.usr_id,
            uf.usf_name_intern,
            ud.usd_value
        FROM {table_prefix}users u
        JOIN {table_prefix}user_data ud ON u.usr_id = ud.usd_usr_id
        JOIN {table_prefix}user_fields uf ON ud.usd_usf_id = uf.usf_id
        WHERE uf.usf_id IN ({",".join(str(fid) for fid in field_ids)})
          AND u.usr_valid = 1
        ORDER BY u.usr_id
    """

    with pymysql.connect(
        host=host,
        port=port,
        database=database,
        user=user,
        password=password,
        cursorclass=pymysql.cursors.DictCursor,
    ) as conn:
        with conn.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()

    members_by_id: dict[int, MemberRecord] = {}
    for row in rows:
        usr_id = row["usr_id"]
        field_name = row["usf_name_intern"]
        value = row["usd_value"]

        if usr_id not in members_by_id:
            members_by_id[usr_id] = MemberRecord()

        if field_name == "SIPPE" and value:
            value = sippe_mapping.get(value, value)

        attr_name = FIELD_TO_ATTR.get(field_name)
        if attr_name:
            setattr(members_by_id[usr_id], attr_name, value)

    return list(members_by_id.values())


def export_to_excel(members: list[MemberRecord], output_path: Path) -> None:
    """Export member records to styled Excel file."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sorted_members = sorted(members, key=lambda m: (m.sippe or "", m.nachname or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Mitglieder"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    data_alignment = Alignment(vertical="center")
    alt_row_fill = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
    )

    ws.append(COLUMN_HEADERS)

    for col_num, _ in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_num, member in enumerate(sorted_members, 2):
        row_data = [
            member.mitglieds_nr,
            member.nachname,
            member.vorname,
            member.kontoinhaber,
            member.familien_nr,
            member.sippe,
        ]
        ws.append(row_data)

        for col_num in range(1, len(COLUMN_HEADERS) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = data_alignment
            if row_num % 2 == 0:
                cell.fill = alt_row_fill

    column_widths = [15, 18, 18, 38, 12, 22]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[1].height = 25

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def generate_filename() -> str:
    """Generate filename with current date and time."""
    now = datetime.now().strftime("%Y-%m-%d_%H%M")
    return f"{now}_Mitgliederliste.xlsx"


def upload_to_admidio(
    local_file: Path,
    remote_filename: str,
    ssh_host: str,
    ssh_user: str,
    ssh_key_path: str,
    db_host: str,
    db_port: int,
    db_name: str,
    db_user: str,
    db_password: str,
    table_prefix: str,
) -> None:
    """Upload file to Admidio and register in database."""
    expanded_key_path = Path(ssh_key_path).expanduser()
    remote_path = f"{ADMIDIO_VOLUME_PATH}/{remote_filename}"

    scp_cmd = [
        "scp",
        "-i",
        str(expanded_key_path),
        "-o",
        "StrictHostKeyChecking=no",
        str(local_file),
        f"{ssh_user}@{ssh_host}:{remote_path}",
    ]
    subprocess.run(scp_cmd, check=True)

    chown_cmd = [
        "ssh",
        "-i",
        str(expanded_key_path),
        "-o",
        "StrictHostKeyChecking=no",
        f"{ssh_user}@{ssh_host}",
        f"chown www-data:www-data '{remote_path}'",
    ]
    subprocess.run(chown_cmd, check=True)

    file_uuid = str(uuid.uuid4())

    with pymysql.connect(
        host=db_host,
        port=db_port,
        database=db_name,
        user=db_user,
        password=db_password,
        cursorclass=pymysql.cursors.DictCursor,
    ) as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT fil_id FROM {table_prefix}files 
                WHERE fil_fol_id = %s AND fil_name = %s
                """,
                (ADMIDIO_MITGLIEDERLISTE_FOLDER_ID, remote_filename),
            )
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    f"""
                    UPDATE {table_prefix}files 
                    SET fil_timestamp = NOW()
                    WHERE fil_id = %s
                    """,
                    (existing["fil_id"],),
                )
            else:
                cursor.execute(
                    f"""
                    INSERT INTO {table_prefix}files 
                    (fil_fol_id, fil_uuid, fil_name, fil_locked, fil_counter, fil_usr_id, fil_timestamp)
                    VALUES (%s, %s, %s, 0, 0, %s, NOW())
                    """,
                    (
                        ADMIDIO_MITGLIEDERLISTE_FOLDER_ID,
                        file_uuid,
                        remote_filename,
                        ADMIDIO_SYSTEM_USER_ID,
                    ),
                )

        conn.commit()


def main() -> None:
    """Main entry point for the member data pipeline."""
    secrets = load_secrets(SECRETS_FILE)

    ssh_host = secrets["HETZNER_SSH_HOST"]
    ssh_user = secrets["HETZNER_SSH_USER"]
    ssh_key_path = secrets["HETZNER_SSH_KEY_PATH"]

    db_name = secrets["ADMIDIO_DB_NAME"]
    db_user = secrets["ADMIDIO_DB_USER"]
    db_password = secrets["ADMIDIO_DB_PASSWORD"]
    table_prefix = secrets["ADMIDIO_TABLE_PREFIX"]

    with ssh_tunnel(
        ssh_host=ssh_host,
        ssh_user=ssh_user,
        ssh_key_path=ssh_key_path,
        local_port=SSH_TUNNEL_LOCAL_PORT,
        remote_host=DOCKER_DB_HOST,
        remote_port=SSH_TUNNEL_REMOTE_PORT,
    ):
        sippe_mapping = fetch_sippe_mapping(
            host="127.0.0.1",
            port=SSH_TUNNEL_LOCAL_PORT,
            database=db_name,
            user=db_user,
            password=db_password,
            table_prefix=table_prefix,
        )

        members = fetch_member_data(
            host="127.0.0.1",
            port=SSH_TUNNEL_LOCAL_PORT,
            database=db_name,
            user=db_user,
            password=db_password,
            table_prefix=table_prefix,
            sippe_mapping=sippe_mapping,
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            filename = generate_filename()
            local_file = Path(tmpdir) / filename

            export_to_excel(members, local_file)

            upload_to_admidio(
                local_file=local_file,
                remote_filename=filename,
                ssh_host=ssh_host,
                ssh_user=ssh_user,
                ssh_key_path=ssh_key_path,
                db_host="127.0.0.1",
                db_port=SSH_TUNNEL_LOCAL_PORT,
                db_name=db_name,
                db_user=db_user,
                db_password=db_password,
                table_prefix=table_prefix,
            )

    print(f"Uploaded {len(members)} members to Admidio as {filename}")


if __name__ == "__main__":
    main()
